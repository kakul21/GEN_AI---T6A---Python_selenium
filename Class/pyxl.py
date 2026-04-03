import pytest
import openpyxl

wb = openpyxl.Workbook()
sheetName = "sheet1"

if sheetName in wb.sheetnames:
    ws = wb[sheetName]
else:
    ws = wb.create_sheet(sheetName)

ws["A1"]="Username"
ws["B1"]="Password"
# wb.save("sample.xlsx")
ws.append(["error_user","secret_sauce"])
ws.append(["performance_glitch_user","secret_sauce"])
ws.append(["problem_user","secret_sauce"])
ws.append(["locked_out_user","secret_sauce"])
ws.append(["standard_user","secret_sauce"])

wb.save("sample.xlsx")

## To print all the value each cell
# for row in ws.rows:
#     for cell in row:
#         print(cell.value)


## To print the value present in each row (row wise)
for row in ws.iter_rows(values_only=True):
    print(row)

## To print the value present in each column (column wise)
for col in ws.iter_cols(values_only=True):
    print(col)







